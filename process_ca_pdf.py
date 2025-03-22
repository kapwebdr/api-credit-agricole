import pandas as pd
import os
from dotenv import load_dotenv
import re
import datetime
import json
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
import argparse
import sys
import ca_common

def load_tva_rules():
    """Charge les règles de TVA depuis le fichier JSON"""
    try:
        with open('tva_rules.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Erreur lors du chargement du fichier de règles TVA: {e}")
        # Règles par défaut
        return {
            "tva_rates": {
                "standard": 20.0,
                "intermédiaire": 10.0,
                "réduit": 5.5,
                "particulier": 7.0,
                "exonéré": 0.0
            },
            "keywords": {
                "standard": ["ovh", "amazon"],
                "intermédiaire": ["restaurant", "resto"],
                "réduit": ["alimentation"],
                "particulier": ["rénovation"],
                "exonéré": ["formation", "impôt"]
            }
        }

def determine_tva_type(libelle, rules):
    """Détermine le type de TVA en fonction du libellé et des règles"""
    if not libelle or pd.isna(libelle):
        return "standard"
    
    libelle = str(libelle).lower()
    
    # Parcourir les règles pour trouver une correspondance
    for tva_type, keywords in rules["keywords"].items():
        for keyword in keywords:
            if keyword.lower() in libelle:
                return tva_type
    
    # Par défaut, TVA standard
    return "standard"

def process_ca_statement(input_file=None, output_dir=None):
    """
    Traite un relevé bancaire du Crédit Agricole au format Excel
    """
    # Vérifier que le fichier d'entrée existe
    if not input_file or not os.path.exists(input_file):
        print(f"Erreur: Le fichier d'entrée n'existe pas: {input_file}")
        return None
    
    # Utiliser le même répertoire que le fichier d'entrée si aucun répertoire de sortie n'est spécifié
    if not output_dir:
        output_dir = os.path.dirname(input_file)
        print(f"Aucun répertoire de sortie spécifié, utilisation du répertoire du fichier d'entrée: {output_dir}")
    
    # Créer le répertoire de sortie s'il n'existe pas
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
        print(f"Répertoire de sortie créé: {output_dir}")
    
    print(f"Traitement du fichier: {input_file}")
    print(f"Répertoire de sortie: {output_dir}")
    
    # Charger les règles TVA
    tva_rules = load_tva_rules()
    TVA_RATES = tva_rules["tva_rates"]
    
    try:
        # Méthode 1: Essayer de trouver l'en-tête sur les 30 premières lignes
        header_row = None
        
        # Tenter plusieurs lectures pour trouver l'entête
        for i in range(30):  # Vérifier jusqu'à 30 lignes
            try:
                temp_df = pd.read_excel(input_file, header=i)
                # Vérifier si les colonnes contiennent 'Date', 'Libellé', 'Débit', 'Crédit'
                cols = [str(col).lower() for col in temp_df.columns]
                if ('date' in cols and 
                    any('lib' in col for col in cols) and 
                    any(term in ''.join(cols) for term in ['débit', 'debit']) and 
                    any(term in ''.join(cols) for term in ['crédit', 'credit'])):
                    header_row = i
                    break
            except Exception as e:
                print(f"Erreur lors de la tentative de lecture avec header={i}: {e}")
                continue
        
        # Méthode 2: Si l'en-tête n'est toujours pas trouvé, tenter une approche différente
        if header_row is None:
            print("Tentative de détection alternative de l'en-tête...")
            
            # Lire tout le fichier sans en-tête
            raw_df = pd.read_excel(input_file, header=None)
            
            # Rechercher une ligne qui contient les mots-clés Date, Libellé/Lib et Débit/Crédit
            for idx, row in raw_df.iterrows():
                row_str = ' '.join([str(val).lower() for val in row.values if pd.notna(val)])
                if ('date' in row_str and 
                    any(term in row_str for term in ['libellé', 'libelle', 'lib']) and
                    any(term in row_str for term in ['débit', 'debit', 'crédit', 'credit'])):
                    header_row = idx
                    print(f"En-tête trouvé à la ligne {header_row + 1}")
                    break
        
        if header_row is None:
            print("Impossible de trouver les en-têtes du tableau. Essai de lecture directe...")
            # Tenter de lire directement le fichier avec un format prédéfini
            try:
                # Lire les 50 premières lignes pour diagnostic
                preview_df = pd.read_excel(input_file, header=None, nrows=50)
                print("Aperçu des premières lignes du fichier:")
                for idx, row in preview_df.head(50).iterrows():
                    print(f"Ligne {idx+1}: {row.values.tolist()}")
                
                # Dernier recours: rechercher une ligne qui contient 'Date'
                for idx, row in preview_df.iterrows():
                    row_values = [str(val).strip().lower() for val in row.values if pd.notna(val)]
                    if any('date' in val for val in row_values):
                        header_row = idx
                        print(f"Potentiel en-tête à la ligne {header_row + 1}: {row.values}")
                        break
            except Exception as diag_error:
                print(f"Erreur lors du diagnostic: {diag_error}")
        
        if header_row is None:
            print("Impossible de trouver l'en-tête. Utilisation de header=0 par défaut.")
            header_row = 0
        
        # Lire le fichier avec l'en-tête identifié
        df = pd.read_excel(input_file, header=header_row)
        print(f"Fichier lu avec en-tête à la ligne {header_row + 1}, colonnes trouvées: {df.columns.tolist()}")
        
        # Nettoyer et renommer les colonnes pour une utilisation cohérente
        df.columns = [str(col).strip() for col in df.columns]
        
        # Mapper les noms de colonnes
        column_mapping = {}
        for col in df.columns:
            col_lower = str(col).lower()
            if 'date' in col_lower:
                column_mapping[col] = 'Date'
            elif any(term in col_lower for term in ['libellé', 'libelle', 'lib']):
                column_mapping[col] = 'Libellé'
            elif any(term in col_lower for term in ['débit', 'debit']):
                column_mapping[col] = 'Débit'
            elif any(term in col_lower for term in ['crédit', 'credit']):
                column_mapping[col] = 'Crédit'
        
        # Renommer les colonnes identifiées
        df = df.rename(columns=column_mapping)
        
        # Vérifier que les colonnes essentielles sont présentes
        required_cols = ['Date', 'Libellé', 'Débit', 'Crédit']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            print(f"Colonnes manquantes: {missing_cols}")
            print(f"Colonnes disponibles: {df.columns.tolist()}")
            
            # Tentative de réparation pour colonnes manquantes
            if 'Date' not in df.columns and any('date' in str(col).lower() for col in df.columns):
                date_col = next(col for col in df.columns if 'date' in str(col).lower())
                df = df.rename(columns={date_col: 'Date'})
                missing_cols.remove('Date')
            
            if len(missing_cols) > 0:
                print("Impossible de trouver toutes les colonnes requises. Traitement impossible.")
                return None
        
        # Nettoyer les données
        # Convertir les valeurs vides en NaN
        df = df.replace('', pd.NA)
        
        # Nettoyer la colonne Libellé en remplaçant les retours à la ligne par des espaces
        if 'Libellé' in df.columns:
            df['Libellé'] = df['Libellé'].apply(lambda x: re.sub(r'[\n\r\t]+', ' ', str(x)) if pd.notna(x) else x)
            # Supprimer les espaces multiples
            df['Libellé'] = df['Libellé'].apply(lambda x: re.sub(r'\s+', ' ', str(x)).strip() if pd.notna(x) else x)
        
        # Convertir la colonne de date
        df['Date'] = pd.to_datetime(df['Date'], dayfirst=True, errors='coerce')
        
        # Convertir les colonnes numériques
        for col in ['Débit', 'Crédit']:
            # Remplacer les non-numériques par NaN
            if df[col].dtype == 'object':
                # Remplacer les virgules par des points
                df[col] = df[col].astype(str).str.replace(',', '.', regex=False)
                # Gérer le format "123,45 €" ou "123.45 €"
                df[col] = df[col].apply(lambda x: re.sub(r'[^\d.,]', '', str(x)) if pd.notna(x) else pd.NA)
                # Gérer le cas où les nombres utilisent la virgule comme séparateur décimal
                df[col] = df[col].apply(lambda x: x.replace(',', '.') if isinstance(x, str) and ',' in x else x)
            df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Remplacer les NaN par des zéros
        df['Débit'] = df['Débit'].fillna(0)
        df['Crédit'] = df['Crédit'].fillna(0)
        
        # Créer une colonne Montant (positif pour les crédits, négatif pour les débits)
        df['Montant'] = df['Crédit'] - df['Débit']
        
        # Filtrer les lignes sans date ou sans montant
        df = df[df['Date'].notna() & ((df['Débit'] > 0) | (df['Crédit'] > 0))]
        
        # Afficher quelques lignes pour diagnostic
        print(f"Échantillon de données traitées:")
        print(df.head())
        
        # Séparer crédits et débits
        credits = df[df['Montant'] > 0].copy()
        debits = df[df['Montant'] < 0].copy()
        debits['Montant'] = debits['Montant'].abs()  # Valeur absolue pour les débits
        
        # Ajouter les colonnes pour le type de TVA sans calculer les valeurs (seront calculées dans Excel)
        for df_op in [credits, debits]:
            # Ajouter une colonne pour le type de TVA
            df_op['Type TVA'] = df_op['Libellé'].apply(lambda x: determine_tva_type(x, tva_rules))
            # Ajouter une colonne pour le taux de TVA (ne pas calculer, juste mettre le taux)
            df_op['Taux TVA'] = df_op['Type TVA'].apply(lambda x: TVA_RATES.get(x, 20.0))
        
        # Créer le fichier de sortie
        today = datetime.datetime.now().strftime("%Y%m%d")
        output_file = os.path.join(output_dir, f"ca_operations_{today}.xlsx")
        
        # Utiliser ExcelWriter avec le moteur openpyxl pour les formules
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Écrire les données brutes
            credits.to_excel(writer, sheet_name='Recettes', index=False)
            debits.to_excel(writer, sheet_name='Dépenses', index=False)
            
            # Accéder aux feuilles de calcul
            workbook = writer.book
            
            # Formules pour les feuilles Recettes et Dépenses
            for sheet_name in ['Recettes', 'Dépenses']:
                worksheet = workbook[sheet_name]
                
                # Trouver les colonnes clés
                header_row = 1  # Excel est 1-indexed
                montant_col = None
                taux_tva_col = None
                
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=header_row, column=col).value
                    if cell_value == 'Montant':
                        montant_col = col
                    elif cell_value == 'Taux TVA':
                        taux_tva_col = col
                
                if montant_col and taux_tva_col:
                    # Ajouter des colonnes pour les calculs
                    ht_col = worksheet.max_column + 1
                    tva_col = worksheet.max_column + 2
                    
                    # Ajouter les en-têtes
                    worksheet.cell(row=header_row, column=ht_col, value='Montant HT')
                    worksheet.cell(row=header_row, column=tva_col, value='TVA')
                    
                    # Format monétaire avec 2 décimales et symbole €
                    euro_format = '# ##0.00 €;-# ##0.00 €'
                    
                    # Ajouter les formules
                    for row in range(2, worksheet.max_row + 1):  # Commencer après l'en-tête
                        # Formule pour le montant HT: =MONTANT / (1 + TAUX_TVA/100)
                        montant_cell = worksheet.cell(row=row, column=montant_col).coordinate
                        taux_cell = worksheet.cell(row=row, column=taux_tva_col).coordinate
                        
                        ht_formula = f"={montant_cell}/(1+{taux_cell}/100)"
                        tva_formula = f"={montant_cell}-{worksheet.cell(row=row, column=ht_col).coordinate}"
                        
                        worksheet.cell(row=row, column=ht_col, value=ht_formula)
                        worksheet.cell(row=row, column=tva_col, value=tva_formula)
                        
                        # Appliquer le format monétaire aux colonnes de montants
                        worksheet.cell(row=row, column=montant_col).number_format = euro_format
                        worksheet.cell(row=row, column=ht_col).number_format = euro_format
                        worksheet.cell(row=row, column=tva_col).number_format = euro_format
            
            # Création d'une feuille pour les règles TVA
            rules_sheet = workbook.create_sheet("Règles TVA")
            
            # Ajouter les en-têtes
            rules_sheet.cell(row=1, column=1, value="Type TVA")
            rules_sheet.cell(row=1, column=2, value="Taux")
            rules_sheet.cell(row=1, column=3, value="Mots-clés")
            
            # Ajouter les règles
            row = 2
            for tva_type, rate in TVA_RATES.items():
                rules_sheet.cell(row=row, column=1, value=tva_type)
                rules_sheet.cell(row=row, column=2, value=rate)
                
                # Ajouter les mots-clés s'ils existent
                keywords = tva_rules["keywords"].get(tva_type, [])
                rules_sheet.cell(row=row, column=3, value=", ".join(keywords))
                
                row += 1
            
            # Créer une feuille de résumé avec des formules
            summary_sheet = workbook.create_sheet("Résumé")
            
            # En-têtes du résumé
            summary_sheet.cell(row=1, column=1, value="Type")
            summary_sheet.cell(row=1, column=2, value="Montant TTC")
            summary_sheet.cell(row=1, column=3, value="Montant HT")
            summary_sheet.cell(row=1, column=4, value="TVA")
            
            # Style pour l'en-tête
            header_style = Font(bold=True)
            for col in range(1, 5):
                summary_sheet.cell(row=1, column=col).font = header_style

            # Lignes du résumé
            summary_sheet.cell(row=2, column=1, value="Recettes")
            summary_sheet.cell(row=3, column=1, value="Dépenses")
            summary_sheet.cell(row=4, column=1, value="Solde")
            
            # Références aux feuilles pour les sommes
            # Formules de somme pour Recettes
            recettes_sheet = workbook["Recettes"]
            montant_col_rec = next(col for col in range(1, recettes_sheet.max_column + 1) 
                              if recettes_sheet.cell(row=1, column=col).value == 'Montant')
            ht_col_rec = next(col for col in range(1, recettes_sheet.max_column + 1) 
                        if recettes_sheet.cell(row=1, column=col).value == 'Montant HT')
            tva_col_rec = next(col for col in range(1, recettes_sheet.max_column + 1) 
                         if recettes_sheet.cell(row=1, column=col).value == 'TVA')
            taux_tva_col_rec = next(col for col in range(1, recettes_sheet.max_column + 1)
                              if recettes_sheet.cell(row=1, column=col).value == 'Taux TVA')
            
            # Formules de somme pour Dépenses
            depenses_sheet = workbook["Dépenses"]
            montant_col_dep = next(col for col in range(1, depenses_sheet.max_column + 1) 
                               if depenses_sheet.cell(row=1, column=col).value == 'Montant')
            ht_col_dep = next(col for col in range(1, depenses_sheet.max_column + 1) 
                         if depenses_sheet.cell(row=1, column=col).value == 'Montant HT')
            tva_col_dep = next(col for col in range(1, depenses_sheet.max_column + 1) 
                          if depenses_sheet.cell(row=1, column=col).value == 'TVA')
            taux_tva_col_dep = next(col for col in range(1, depenses_sheet.max_column + 1)
                               if depenses_sheet.cell(row=1, column=col).value == 'Taux TVA')
            
            # Formules pour le résumé
            # Recettes
            summary_sheet.cell(row=2, column=2, value=f"=SUM(Recettes!{get_column_letter(montant_col_rec)}2:{get_column_letter(montant_col_rec)}{recettes_sheet.max_row})")
            summary_sheet.cell(row=2, column=3, value=f"=SUM(Recettes!{get_column_letter(ht_col_rec)}2:{get_column_letter(ht_col_rec)}{recettes_sheet.max_row})")
            summary_sheet.cell(row=2, column=4, value=f"=SUM(Recettes!{get_column_letter(tva_col_rec)}2:{get_column_letter(tva_col_rec)}{recettes_sheet.max_row})")
            
            # Dépenses
            summary_sheet.cell(row=3, column=2, value=f"=SUM(Dépenses!{get_column_letter(montant_col_dep)}2:{get_column_letter(montant_col_dep)}{depenses_sheet.max_row})")
            summary_sheet.cell(row=3, column=3, value=f"=SUM(Dépenses!{get_column_letter(ht_col_dep)}2:{get_column_letter(ht_col_dep)}{depenses_sheet.max_row})")
            summary_sheet.cell(row=3, column=4, value=f"=SUM(Dépenses!{get_column_letter(tva_col_dep)}2:{get_column_letter(tva_col_dep)}{depenses_sheet.max_row})")
            
            # Solde
            summary_sheet.cell(row=4, column=2, value="=B2-B3")
            summary_sheet.cell(row=4, column=3, value="=C2-C3")
            summary_sheet.cell(row=4, column=4, value="=D2-D3")
            
            # Ajouter une ligne pour les recettes assujetties à TVA
            summary_sheet.cell(row=5, column=1, value="Recettes assujetties à TVA")
            
            # Formule SUMIF pour filtrer les recettes avec un taux de TVA > 0
            taux_tva_col_letter_rec = get_column_letter(taux_tva_col_rec)
            ht_col_letter_rec = get_column_letter(ht_col_rec)
            tva_col_letter_rec = get_column_letter(tva_col_rec)
            
            # Formule pour les recettes HT avec TVA > 0
            summary_sheet.cell(row=5, column=3, value=f'=SUMIF(Recettes!{taux_tva_col_letter_rec}2:{taux_tva_col_letter_rec}{recettes_sheet.max_row},">0",Recettes!{ht_col_letter_rec}2:{ht_col_letter_rec}{recettes_sheet.max_row})')
            
            # Formule pour la TVA correspondante
            summary_sheet.cell(row=5, column=4, value=f'=SUMIF(Recettes!{taux_tva_col_letter_rec}2:{taux_tva_col_letter_rec}{recettes_sheet.max_row},">0",Recettes!{tva_col_letter_rec}2:{tva_col_letter_rec}{recettes_sheet.max_row})')
            
            # Formule pour le montant TTC
            summary_sheet.cell(row=5, column=2, value="=C5+D5")
            
            # Appliquer le style de nombre entier pour les colonnes TVA (sans décimales)
            integer_format = numbers.FORMAT_NUMBER_00.replace('.00', '')
            
            # Appliquer le format sans décimales à toutes les cellules numériques de l'onglet Résumé
            for row in range(2, 6):  # Ajusté pour correspondre aux lignes actuelles
                for col in range(2, 5):  # Colonnes B, C, D (Montant TTC, Montant HT, TVA)
                    summary_sheet.cell(row=row, column=col).number_format = integer_format
            
            # Appliquer un fond de couleur aux cellules importantes pour l'import
            highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Jaune
            
            # Cellules à mettre en évidence:
            # 1. TVA des recettes assujetties
            summary_sheet.cell(row=5, column=4).fill = highlight_fill
            # 2. HT des recettes assujetties
            summary_sheet.cell(row=5, column=3).fill = highlight_fill
            # 3. TVA des dépenses (total)
            summary_sheet.cell(row=3, column=4).fill = highlight_fill
            
            # Créer une feuille de synthèse TVA avec des formules
            tva_sheet = workbook.create_sheet("Synthèse TVA")
            
            # En-têtes
            tva_sheet.cell(row=1, column=1, value="Type TVA")
            tva_sheet.cell(row=1, column=2, value="Taux")
            tva_sheet.cell(row=1, column=3, value="TVA Collectée")
            tva_sheet.cell(row=1, column=4, value="TVA Déductible")
            tva_sheet.cell(row=1, column=5, value="Solde TVA")
            
            # Ajouter une ligne pour chaque taux de TVA
            row = 2
            for tva_type, rate in TVA_RATES.items():
                tva_sheet.cell(row=row, column=1, value=tva_type)
                tva_sheet.cell(row=row, column=2, value=f"{rate}%")
                
                # Formules pour TVA collectée (recettes) et déductible (dépenses)
                # Filtre par type de TVA
                tva_sheet.cell(row=row, column=3, value=f'=SUMIF(Recettes!TypeTVA,"={tva_type}",Recettes!TVA)')
                tva_sheet.cell(row=row, column=4, value=f'=SUMIF(Dépenses!TypeTVA,"={tva_type}",Dépenses!TVA)')
                tva_sheet.cell(row=row, column=5, value=f"=C{row}-D{row}")
                
                row += 1
            
            # Ajouter une ligne de total
            tva_sheet.cell(row=row, column=1, value="TOTAL")
            tva_sheet.cell(row=row, column=3, value=f"=SUM(C2:C{row-1})")
            tva_sheet.cell(row=row, column=4, value=f"=SUM(D2:D{row-1})")
            tva_sheet.cell(row=row, column=5, value=f"=C{row}-D{row}")
        
        # Après avoir sauvegardé, rouvrir et corriger les formules SUMIF et appliquer les formats
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        
        # Définir les formats de nombre
        integer_format = numbers.FORMAT_NUMBER_00.replace('.00', '')
        euro_format = '# ##0.00 €;-# ##0.00 €'
        
        # Accéder à la feuille de synthèse TVA
        tva_sheet = wb["Synthèse TVA"]
        
        # Récupérer les indices de colonnes
        recettes_sheet = wb["Recettes"]
        depenses_sheet = wb["Dépenses"]
        summary_sheet = wb["Résumé"]
        
        # Appliquer le format monétaire aux colonnes TVA de la synthèse TVA (première partie)
        for row in range(2, tva_sheet.max_row + 1):
            tva_sheet.cell(row=row, column=3).number_format = euro_format
            tva_sheet.cell(row=row, column=4).number_format = euro_format
            tva_sheet.cell(row=row, column=5).number_format = euro_format
        
        # Trouver les colonnes Type TVA et TVA dans les feuilles
        type_tva_col_rec = get_column_letter(next(i+1 for i, cell in enumerate(recettes_sheet[1]) if cell.value == "Type TVA"))
        tva_col_rec = get_column_letter(next(i+1 for i, cell in enumerate(recettes_sheet[1]) if cell.value == "TVA"))
        
        type_tva_col_dep = get_column_letter(next(i+1 for i, cell in enumerate(depenses_sheet[1]) if cell.value == "Type TVA"))
        tva_col_dep = get_column_letter(next(i+1 for i, cell in enumerate(depenses_sheet[1]) if cell.value == "TVA"))
        
        # Corriger les formules SUMIF
        for row in range(2, tva_sheet.max_row):
            tva_type = tva_sheet.cell(row=row, column=1).value
            if tva_type in TVA_RATES:
                # Corriger les formules avec les bonnes références de colonnes
                tva_sheet.cell(row=row, column=3).value = f'=SUMIF(Recettes!{type_tva_col_rec}2:{type_tva_col_rec}{recettes_sheet.max_row},"{tva_type}",Recettes!{tva_col_rec}2:{tva_col_rec}{recettes_sheet.max_row})'
                tva_sheet.cell(row=row, column=4).value = f'=SUMIF(Dépenses!{type_tva_col_dep}2:{type_tva_col_dep}{depenses_sheet.max_row},"{tva_type}",Dépenses!{tva_col_dep}2:{tva_col_dep}{depenses_sheet.max_row})'
        
        # Ajouter une synthèse globale TTC
        # Laisser quelques lignes vides après le tableau existant
        start_row = tva_sheet.max_row + 3
        
        # Titre de la section
        tva_sheet.cell(row=start_row, column=1, value="SYNTHÈSE GLOBALE").font = Font(bold=True, size=12)
        tva_sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)
        tva_sheet.cell(row=start_row, column=1).alignment = Alignment(horizontal='center')
        
        # En-têtes de la synthèse globale
        tva_sheet.cell(row=start_row+2, column=1, value="Type").font = Font(bold=True)
        tva_sheet.cell(row=start_row+2, column=2, value="Montant TTC").font = Font(bold=True)
        tva_sheet.cell(row=start_row+2, column=3, value="Montant HT").font = Font(bold=True)
        tva_sheet.cell(row=start_row+2, column=4, value="TVA").font = Font(bold=True)
        
        # Récupérer les références aux cellules du résumé
        # Recettes
        tva_sheet.cell(row=start_row+3, column=1, value="Total Recettes")
        tva_sheet.cell(row=start_row+3, column=2, value=f"=Résumé!B2")  # Montant TTC
        tva_sheet.cell(row=start_row+3, column=3, value=f"=Résumé!C2")  # Montant HT
        tva_sheet.cell(row=start_row+3, column=4, value=f"=Résumé!D2")  # TVA
        
        # Dépenses
        tva_sheet.cell(row=start_row+4, column=1, value="Total Dépenses")
        tva_sheet.cell(row=start_row+4, column=2, value=f"=Résumé!B3")  # Montant TTC
        tva_sheet.cell(row=start_row+4, column=3, value=f"=Résumé!C3")  # Montant HT
        tva_sheet.cell(row=start_row+4, column=4, value=f"=Résumé!D3")  # TVA
        
        # Solde
        tva_sheet.cell(row=start_row+5, column=1, value="Solde").font = Font(bold=True)
        tva_sheet.cell(row=start_row+5, column=2, value=f"=B{start_row+3}-B{start_row+4}")  # Montant TTC
        tva_sheet.cell(row=start_row+5, column=3, value=f"=C{start_row+3}-C{start_row+4}")  # Montant HT
        tva_sheet.cell(row=start_row+5, column=4, value=f"=D{start_row+3}-D{start_row+4}")  # TVA
        
        # Ajouter les recettes assujetties (pour compléter la synthèse)
        tva_sheet.cell(row=start_row+7, column=1, value="Recettes assujetties à TVA")
        tva_sheet.cell(row=start_row+7, column=2, value=f"=Résumé!B5")  # Montant TTC
        tva_sheet.cell(row=start_row+7, column=3, value=f"=Résumé!C5")  # Montant HT
        tva_sheet.cell(row=start_row+7, column=4, value=f"=Résumé!D5")  # TVA
        
        # Mise en forme des valeurs numériques dans la synthèse globale (format monétaire)
        for row in range(start_row+3, start_row+8):
            for col in range(2, 5):
                if tva_sheet.cell(row=row, column=col).value:
                    tva_sheet.cell(row=row, column=col).number_format = euro_format
        
        # S'assurer que l'onglet Résumé garde bien le format entier sans décimales
        for row in range(2, 6):  # Ajusté pour correspondre aux lignes actuelles
            for col in range(2, 5):  # Colonnes B, C, D (Montant TTC, Montant HT, TVA)
                summary_sheet.cell(row=row, column=col).number_format = integer_format
                
        # Mettre en évidence les cellules importantes avec un fond coloré
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Jaune
        
        # Mettre en évidence les mêmes cellules que dans l'onglet Résumé
        tva_sheet.cell(row=start_row+7, column=4).fill = highlight_fill  # TVA des recettes assujetties
        tva_sheet.cell(row=start_row+7, column=3).fill = highlight_fill  # HT des recettes assujetties
        tva_sheet.cell(row=start_row+4, column=4).fill = highlight_fill  # TVA des dépenses totales
        
        # Ajuster automatiquement la largeur des colonnes pour toutes les feuilles
        for sheet_name in wb.sheetnames:
            adjust_column_width(wb[sheet_name])
        
        # Sauvegarder les modifications
        wb.save(output_file)
        
        print(f"Traitement terminé. Fichier généré: {output_file}")
        return output_file
            
    except Exception as e:
        print(f"Erreur lors du traitement: {e}")
        import traceback
        traceback.print_exc()
        return None

def get_column_letter(column_index):
    """Convertit un indice de colonne en lettre (1 = A, 2 = B, etc.)"""
    from openpyxl.utils import get_column_letter as openpyxl_get_column_letter
    return openpyxl_get_column_letter(column_index)

def adjust_column_width(worksheet):
    """Ajuste la largeur des colonnes en fonction du contenu"""
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Obtenir la lettre de la colonne
        
        # Trouver la longueur maximale dans la colonne
        for cell in col:
            if cell.value:
                # Longueur du contenu + marge
                cell_length = len(str(cell.value)) + 2
                if cell_length > max_length:
                    max_length = cell_length
        
        # Définir une largeur minimale
        if max_length < 10:
            max_length = 10
        
        # Définir une largeur maximale
        if max_length > 50:
            max_length = 50
            
        # Ajuster la largeur
        worksheet.column_dimensions[column].width = max_length

def process_files_automatically():
    """
    Traite automatiquement tous les fichiers téléchargés pour le mois précédent
    """
    # Obtenir le répertoire dynamique pour le mois précédent
    dynamic_dir = ca_common.get_dynamic_directory()
    
    # Rechercher les fichiers de compte
    account_files = ca_common.get_account_files(dynamic_dir)
    
    if not account_files:
        print(f"Aucun fichier à traiter trouvé dans {dynamic_dir}")
        return False
    
    print(f"Fichiers trouvés: {len(account_files)}")
    
    # Traiter chaque fichier
    success_count = 0
    for filepath, account in account_files:
        print(f"\n--- Traitement du fichier pour le compte {account} ---")
        result = process_ca_statement(filepath, dynamic_dir)
        if result:
            success_count += 1
            print(f"Traitement réussi pour {account}")
        else:
            print(f"Échec du traitement pour {account}")
    
    print(f"\nTraitement terminé: {success_count}/{len(account_files)} fichiers traités avec succès")
    
    return success_count > 0

def main():
    # Analyse des arguments en ligne de commande
    parser = argparse.ArgumentParser(description="Traitement des relevés bancaires Crédit Agricole")
    parser.add_argument("--input", "-i", help="Chemin du fichier Excel à traiter")
    parser.add_argument("--output", "-o", help="Répertoire de sortie pour le fichier résultat")
    parser.add_argument("--env", help="Chemin vers un fichier .env alternatif")
    parser.add_argument("--account", help="Numéro de compte spécifique à traiter")
    args = parser.parse_args()
    
    # Charger le fichier .env approprié
    ca_common.load_environment(args.env)
    
    # Mode spécifique si un fichier d'entrée est fourni
    if args.input:
        # Traiter un fichier spécifique
        result = process_ca_statement(args.input, args.output)
        
        if not result:
            print("Le traitement a échoué.")
            sys.exit(1)
        else:
            print(f"Le traitement a réussi. Fichier généré: {result}")
            sys.exit(0)
    elif args.account:
        # Traiter un compte spécifique en mode automatique
        dynamic_dir = ca_common.get_dynamic_directory()
        account_files = ca_common.get_account_files(dynamic_dir, args.account)
        
        if not account_files:
            print(f"Aucun fichier trouvé pour le compte {args.account}")
            sys.exit(1)
        
        filepath, _ = account_files[0]
        result = process_ca_statement(filepath, args.output or dynamic_dir)
        
        if not result:
            print("Le traitement a échoué.")
            sys.exit(1)
        else:
            print(f"Le traitement a réussi. Fichier généré: {result}")
            sys.exit(0)
    else:
        # Mode automatique - traiter tous les fichiers pour le mois précédent
        print("Mode automatique - traitement de tous les fichiers du mois précédent")
        success = process_files_automatically()
        
        if not success:
            print("Le traitement automatique a échoué.")
            sys.exit(1)
        else:
            print("Le traitement automatique a réussi.")
            sys.exit(0)

if __name__ == "__main__":
    main() 